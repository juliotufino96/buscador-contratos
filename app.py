import streamlit as st
import pandas as pd
import os
import glob
import re
import unicodedata
import io
import gdown
import datetime
from PIL import Image  # NUEVO: Librería para manejar imágenes
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ================= CONFIGURACIÓN =================
# Intentamos cargar la imagen local para el ícono de la pestaña
try:
    icono = Image.open("logo.jpg") # Asegúrate de que se llame así en tu GitHub
except FileNotFoundError:
    icono = "🏢" # Respaldo por si la imagen no se encuentra

st.set_page_config(page_title="Buscador de Contratos", page_icon=icono, layout="centered")

# ID de tu carpeta pública de Drive
FOLDER_ID = "1IenfFVfGPxVyEjBaK7M_1JtAKbBGaWlf"
RUTA_BASE = "Datos_Descargados"
ruta_parquet = os.path.join(RUTA_BASE, "historico.parquet")

# ================= FUNCIONES DE LIMPIEZA =================
def normalizar_para_busqueda(texto):
    if pd.isna(texto): return ""
    texto = str(texto).lower()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('utf-8')
    return re.sub(r'[.,]', '', texto).strip()

def limpiar_y_agrupar_proveedor(texto):
    if pd.isna(texto): return ""
    texto = str(texto).upper()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('utf-8')
    texto = re.sub(r'[.,-]', ' ', texto) 
    texto = re.sub(r'\s+', ' ', texto).strip()
    
    patron_legal = r'\b(S\s*A\s*DE\s*C\s*V|S\s*DE\s*R\s*L\s*DE\s*C\s*V|S\s*DE\s*R\s*L|S\s*A\s*P\s*I\s*DE\s*C\s*V|S\s*A\s*B\s*DE\s*C\s*V|S\s*A|S\s*C|R\s*L\s*DE\s*C\s*V|S\s*P\s*R\s*DE\s*R\s*L|A\s*C)\b'
    texto = re.sub(patron_legal, '', texto).strip()
    
    return re.sub(r'\s+', ' ', texto).strip()

def limpiar_institucion(texto):
    if pd.isna(texto): return ""
    texto = str(texto).upper()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('utf-8')
    texto = re.sub(r'[.,]', '', texto)
    return texto.strip()

def leer_csv_seguro(ruta):
    codificaciones = ['utf-8-sig', 'latin-1', 'windows-1252']
    for cod in codificaciones:
        try:
            return pd.read_csv(ruta, encoding=cod, dtype=str)
        except Exception:
            continue
    return None

def create_pivot_with_subtotals(df, values_col, agg_func):
    base = pd.pivot_table(df, values=values_col, index=['Orden de gobierno', 'Ramo_Sort', 'Clave Ramo', 'Institución'], columns=['Año'], aggfunc=agg_func, fill_value=0).reset_index()
    base['Is_Subtotal'], base['Is_Total'] = False, False

    sub = pd.pivot_table(df, values=values_col, index=['Orden de gobierno'], columns=['Año'], aggfunc=agg_func, fill_value=0).reset_index()
    sub['Ramo_Sort'] = 99999
    sub['Clave Ramo'] = 'Total ' + sub['Orden de gobierno']
    sub['Institución'] = ''
    sub['Is_Subtotal'], sub['Is_Total'] = True, False

    tot = pd.pivot_table(df, values=values_col, index=lambda x: 'Total General', columns=['Año'], aggfunc=agg_func, fill_value=0).reset_index()
    tot.rename(columns={'index': 'Orden de gobierno'}, inplace=True)
    tot['Orden de gobierno'] = 'Total General'
    tot['Ramo_Sort'] = 999999 
    tot['Clave Ramo'] = ''
    tot['Institución'] = ''
    tot['Is_Subtotal'], tot['Is_Total'] = False, True

    res = pd.concat([base, sub, tot], ignore_index=True).fillna(0)
    res['Orden_Sort'] = res['Orden de gobierno'].apply(lambda x: 'ZZZZ' if x == 'Total General' else x)
    res.sort_values(by=['Orden_Sort', 'Ramo_Sort', 'Clave Ramo', 'Institución'], inplace=True)
    res.drop(columns=['Orden_Sort'], inplace=True)
    return res

def formatear_y_escribir_tabla_integrada(ws, pt_count, pt_sum, nombre_proveedor, start_row=2):
    guinda_fill = PatternFill(start_color='9B2247', end_color='9B2247', fill_type='solid')
    verde_fill = PatternFill(start_color='1E5B4F', end_color='1E5B4F', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_thick = Side(style='thick')
    border_thin = Side(style='thin')

    columnas_periodos = [c for c in pt_count.columns if str(c).isdigit() or str(c) == 'Sin Año']
    num_rows, total_columnas = len(pt_count), 3 + len(columnas_periodos) * 2
    row_titulo, row_super_header, row_sub_header = start_row, start_row + 1, start_row + 2

    ws.cell(row=row_titulo, column=1, value=f"{nombre_proveedor}")
    ws.merge_cells(start_row=row_titulo, start_column=1, end_row=row_titulo, end_column=total_columnas)
    for c in range(1, total_columnas + 1):
        cell = ws.cell(row=row_titulo, column=c)
        cell.fill, cell.font, cell.alignment = guinda_fill, white_font, align_center

    for c_idx, nombre_col in enumerate(["Orden de gobierno", "Clave Ramo", "Institución"], 1):
        ws.cell(row=row_super_header, column=c_idx, value=nombre_col)
        ws.merge_cells(start_row=row_super_header, start_column=c_idx, end_row=row_sub_header, end_column=c_idx)
        for r_idx in [row_super_header, row_sub_header]:
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.fill, cell.font, cell.alignment = verde_fill, white_font, align_center

    col_actual = 4
    for periodo in columnas_periodos:
        ws.cell(row=row_super_header, column=col_actual, value=periodo)
        ws.merge_cells(start_row=row_super_header, start_column=col_actual, end_row=row_super_header, end_column=col_actual + 1)
        for c in range(col_actual, col_actual + 2):
            cell = ws.cell(row=row_super_header, column=c)
            cell.fill, cell.font, cell.alignment = guinda_fill, white_font, align_center
        for idx, sub_val in enumerate(["Número de procedimiento", "Importe (Millones)"]):
            sc = ws.cell(row=row_sub_header, column=col_actual + idx, value=sub_val)
            sc.fill, sc.font, sc.alignment = guinda_fill, white_font, align_center
        col_actual += 2

    data_start_row = row_sub_header + 1
    for i in range(num_rows):
        r_idx = data_start_row + i
        row_count, row_sum = pt_count.iloc[i], pt_sum.iloc[i]
        
        es_total, es_subtotal = row_count.get('Is_Total', False), row_count.get('Is_Subtotal', False)
        c_ord = ws.cell(row=r_idx, column=1, value=row_count['Orden de gobierno'])
        c_ramo = ws.cell(row=r_idx, column=2, value=row_count['Clave Ramo'])
        c_inst = ws.cell(row=r_idx, column=3, value=row_count['Institución'])

        fill_color = guinda_fill if (es_total or es_subtotal) else verde_fill
        for c in [c_ord, c_ramo, c_inst]: c.fill, c.font = fill_color, white_font

        col_data = 4
        for periodo in columnas_periodos:
            c_cnt = ws.cell(row=r_idx, column=col_data, value=row_count[periodo])
            c_mnt = ws.cell(row=r_idx, column=col_data + 1, value=row_sum[periodo])
            c_cnt.number_format, c_mnt.number_format = '#,##0', '"$"#,##0.00'
            if es_total or es_subtotal:
                for c in [c_cnt, c_mnt]: c.fill, c.font = fill_color, white_font
            col_data += 2

    end_row = data_start_row + num_rows - 1
    for r in range(start_row, end_row + 1):
        for c in range(1, total_columnas + 1):
            cell = ws.cell(row=r, column=c)
            t = border_thick if r == start_row or r == data_start_row else border_thin
            b = border_thick if r == row_sub_header or r == end_row else border_thin
            
            if data_start_row <= r < end_row:
                b = border_thin
                row_current = pt_count.iloc[r - data_start_row]
                row_next = pt_count.iloc[r - data_start_row + 1]
                if row_current['Orden de gobierno'] != row_next['Orden de gobierno'] or row_current.get('Is_Subtotal'):
                    b = border_thick
            cell.border = Border(top=t, bottom=b, left=border_thick if c == 1 else border_thin, right=border_thick if c == total_columnas else border_thin)
    return end_row, total_columnas

# ================= CARGA DE DATOS COMPARTIDA (CACHÉ) =================
@st.cache_data(show_spinner=False)
def cargar_y_procesar_datos():
    url = f'https://drive.google.com/drive/folders/{FOLDER_ID}'
    gdown.download_folder(url, output=RUTA_BASE, quiet=True, use_cookies=False)
    
    columnas_finales = [
        "Nombre del archivo", "Orden de gobierno", "Clave Ramo",
        "Siglas de la Institución", "Institución", "Número de procedimiento",
        "Núm. del contrato", "Fecha de inicio del contrato",
        "Fecha de fin del contrato", "Importe",
        "Proveedor o contratista", "Dirección del anuncio"
    ]
    
    df_list = []
    if os.path.exists(ruta_parquet): df_list.append(pd.read_parquet(ruta_parquet))
        
    archivos_csv = sorted(glob.glob(os.path.join(RUTA_BASE, "*.csv")))
    for csv in archivos_csv:
        df_temp = leer_csv_seguro(csv)
        if df_temp is not None:
            df_temp = df_temp.rename(columns={'Importe DRC': 'Importe', 'Número del procedimiento': 'Número de procedimiento', 'Importe del contrato': 'Importe'})
            df_temp['Nombre del archivo'] = os.path.basename(csv)
            df_list.append(df_temp)

    if df_list:
        df_maestro = pd.concat(df_list, ignore_index=True).reindex(columns=columnas_finales)
        df_maestro['Proveedor_Limpio'] = df_maestro['Proveedor o contratista'].apply(normalizar_para_busqueda)
        df_maestro['Proveedor_Agrupado'] = df_maestro['Proveedor o contratista'].apply(limpiar_y_agrupar_proveedor)
        return df_maestro
    return pd.DataFrame()

# ================= INTERFAZ WEB STREAMLIT =================
# 1. Cambiamos la proporción de las columnas (antes [1, 8], ahora [1.5, 7]) para darle más espacio al logo
col_img, col_tit = st.columns([1.5, 7])

with col_img:
    try:
        # 2. Aumentamos el tamaño de 70 a 150 (puedes subirlo a 200 o más si lo necesitas)
        st.image("logo.jpg", width=150) 
    except FileNotFoundError:
        st.markdown("<h1>🏢</h1>", unsafe_allow_html=True) 

with col_tit:
    st.markdown("<br>", unsafe_allow_html=True) # Baja un poco el título
    st.title("Consulta de Contratos")

st.markdown("---")

# SECCIÓN 1: SINCRONIZAR DATOS
st.subheader("1. Actualizar Datos desde la Nube")

if st.button("Actualizar"):
    st.cache_data.clear() # Limpia la memoria caché antes de actualizar
    with st.spinner("Descargando archivos desde Google Drive y procesando..."):
        df_maestro = cargar_y_procesar_datos()
        if not df_maestro.empty:
            st.success(f"¡Base de datos lista! {len(df_maestro):,} registros cargados en memoria compartida.")
        else:
            st.error("No se encontraron archivos en la carpeta.")
else:
    # Carga la base de datos de manera silenciosa si ya existe en caché
    with st.spinner("Cargando datos..."):
        df_maestro = cargar_y_procesar_datos()

st.markdown("---")

# SECCIÓN 2: BUSCAR Y DESCARGAR EXCEL
st.subheader("2. Buscar Proveedor y Generar Reporte")
termino = st.text_input("Ingresa el nombre del proveedor (Mínimo 3 letras):", placeholder="Ej. UNIVERSAL EXPORTS")

if df_maestro.empty:
    st.warning("⚠️ Primero debes hacer clic en 'Actualizar' en la parte superior.")
elif termino:
    if len(termino) < 3:
        st.info("ℹ️ Sigue escribiendo para iniciar la búsqueda...")
    else:
        termino_norm = normalizar_para_busqueda(termino)
        
        mask = df_maestro['Proveedor_Limpio'].str.contains(termino_norm, regex=False, na=False)
        df_coincidencias = df_maestro[mask]
        
        proveedores_unicos = df_coincidencias['Proveedor_Agrupado'].dropna().unique()

        if len(proveedores_unicos) == 0:
            st.error("❌ No se encontraron contratos que coincidan con tu búsqueda.")
        else:
            st.success(f"✅ Se encontraron opciones similares. Puedes seleccionar UNA o VARIAS para unificarlas:")
            
            proveedores_seleccionados = st.multiselect(
                "Proveedores encontrados (Selecciona para generar Excel):",
                sorted(proveedores_unicos)
            )

            if proveedores_seleccionados:
                with st.spinner(f"Integrando datos y preparando Excel..."):
                    df_exportar = df_coincidencias[df_coincidencias['Proveedor_Agrupado'].isin(proveedores_seleccionados)].copy()
                    
                    # === TRUCO DEL DISFRAZ (mode) ===
                    proveedor_estandar = df_exportar['Proveedor o contratista'].mode()[0]
                    df_exportar['Proveedor o contratista'] = proveedor_estandar
                    # ================================
                    
                    df_exportar['Importe'] = pd.to_numeric(df_exportar['Importe'].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)
                    df_exportar.insert(df_exportar.columns.get_loc('Importe') + 1, 'Importe (en millones)', df_exportar['Importe'] / 1000000.0)
                    df_exportar['Año'] = df_exportar['Nombre del archivo'].astype(str).str.extract(r'(\d{4})').fillna('Sin Año')
                    
                    df_exportar['Institución'] = df_exportar['Institución'].apply(limpiar_institucion)
                    df_exportar['Orden de gobierno'] = df_exportar['Orden de gobierno'].fillna('No especificado')
                    df_exportar['Clave Ramo'] = df_exportar['Clave Ramo'].fillna('Vacío')
                    df_exportar['Ramo_Sort'] = df_exportar['Clave Ramo'].apply(lambda x: float(x) if str(x).replace('.','',1).isdigit() else 9999)
                    df_exportar['Cuenta'] = 1

                    pt_count = create_pivot_with_subtotals(df_exportar, 'Cuenta', 'sum')
                    pt_sum = create_pivot_with_subtotals(df_exportar, 'Importe (en millones)', 'sum')

                    for col in ['Proveedor_Limpio', 'Proveedor_Agrupado', 'Ramo_Sort', 'Cuenta']: 
                        df_exportar.drop(columns=[col], errors='ignore', inplace=True)

                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_exportar.to_excel(writer, sheet_name='Detalle Contratos', index=False)
                        ws_base = writer.sheets['Detalle Contratos']
                        idx_imp, idx_mill = df_exportar.columns.get_loc("Importe") + 1, df_exportar.columns.get_loc("Importe (en millones)") + 1
                        for r in range(2, len(df_exportar) + 2):
                            ws_base.cell(row=r, column=idx_imp).number_format = '"$"#,##0.00'
                            ws_base.cell(row=r, column=idx_mill).number_format = '"$"#,##0.00'

                        ws_dinamica = writer.book.create_sheet('Resumen')
                        end_row, col_total = formatear_y_escribir_tabla_integrada(ws_dinamica, pt_count, pt_sum, proveedor_estandar, 2)
                        ws_dinamica.column_dimensions['A'].width = 18 
                        ws_dinamica.column_dimensions['B'].width = 18 
                        ws_dinamica.column_dimensions['C'].width = 45 
                        for col in range(4, col_total + 1): 
                            ws_dinamica.column_dimensions[get_column_letter(col)].width = 24
                        
                        writer.book.move_sheet("Resumen", offset=-1)
                
                nombre_archivo = f"Contratos_{re.sub(r'[\\/*?:<>|]', '', proveedor_estandar)}.xlsx"
                
                st.download_button(
                    label=f"⬇️ Descargar Reporte Consolidado",
                    data=output.getvalue(),
                    file_name=nombre_archivo,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

# ================= DISCLAIMER =================
año_actual = datetime.datetime.now().year
st.markdown("<br><br>", unsafe_allow_html=True)
st.markdown("---")
st.markdown(
    f"<p style='text-align: center; font-size: 11px; color: #666;'>"
    f"Fuente: La información fue generada con base en los datos abiertos de los contratos registrados en la plataforma Compras MX "
    f"durante el periodo 2019-{año_actual}, para consulta pública. Disponibles en: "
    f"<a href='https://comprasmx.buengobierno.gob.mx/datos-abiertos' target='_blank'>https://comprasmx.buengobierno.gob.mx/datos-abiertos</a></p>",
    unsafe_allow_html=True
)
